import datetime
import os
from datetime import date
from io import BytesIO
from typing import Iterable

import certifi
import openpyxl
import requests

HOLIDAY_API_URL = "https://holidays-jp.github.io/api/v1/date.json"


def build_year_template(base_template_path: str, year_template_path: str, year: int):
    if not os.path.exists(base_template_path):
        raise FileNotFoundError("Base template not found")

    wb = openpyxl.load_workbook(base_template_path, data_only=False)

    if "計算用" in wb.sheetnames:
        ws_calc = wb["計算用"]
        # Clear existing A/B columns only
        if ws_calc.max_row:
            for row in range(1, ws_calc.max_row + 1):
                ws_calc[f"A{row}"].value = None
                ws_calc[f"B{row}"].value = None
    else:
        ws_calc = wb.create_sheet("計算用")

    holidays = _fetch_holidays()
    rows = []
    for key, value in holidays.items():
        if key.startswith(f"{year}-"):
            rows.append((key, value))

    rows.sort(key=lambda item: item[0])

    for idx, (day, label) in enumerate(rows, start=1):
        try:
            y, m, d = map(int, day.split("-", 2))
            dt = datetime.date(y, m, d)

            cell = ws_calc[f"A{idx}"]
            cell.value = dt
            # 显示成「2026年1月1日」，但值仍是日期数值
            cell.number_format = 'yyyy"年"m"月"d"日"'
        except ValueError:
            # 兜底：无法解析的情况直接写原字符串
            ws_calc[f"A{idx}"].value = day

        ws_calc[f"B{idx}"].value = label

        # 确保输出目录存在
    output_dir = os.path.dirname(year_template_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    wb.save(year_template_path)


def _fetch_holidays():
    resp = requests.get(HOLIDAY_API_URL, timeout=10, verify=certifi.where())
    resp.raise_for_status()
    data = resp.json() if resp.content else {}
    if not isinstance(data, dict):
        return {}
    return data


def export_kintai_xlsx(template_path: str, year: int, month: int, records: Iterable[dict], employee_name: str = ""):
    wb = openpyxl.load_workbook(template_path, data_only=False)
    target_title = f"{year}.{month:02d}"  # e.g. "2025.08"
    if target_title in wb.sheetnames:
        ws = wb[target_title]
    else:
        # Fallback: rename the template's first sheet to the target title
        ws = wb.active if wb.worksheets else wb.create_sheet()
        ws.title = target_title

    # Baseline date in template (month start)
    ws["B2"].value = date(year, month, 1)
    if employee_name:
        ws["H4"].value = employee_name

    for record in records:
        raw_date = record.get("date")
        if not raw_date:
            continue
        try:
            d = date.fromisoformat(raw_date)
        except ValueError:
            continue
        if d.year != year or d.month != month:
            continue

        row = 7 + (d.day - 1)

        # Write end time (column E)
        end_time = record.get("end_time")
        if end_time:
            hh, mm = map(int, end_time.split(":"))
            ws[f"E{row}"].value = openpyxl.utils.datetime.time(hh, mm)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
